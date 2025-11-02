from __future__ import annotations

import argparse
import json
import logging
import math
import os
import re
import time
import traceback
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# Optional SciPy dependency for interpolation and smoothing
try:
    from scipy import ndimage
    from scipy.interpolate import griddata, RBFInterpolator, CubicSpline, RegularGridInterpolator
    from scipy.stats import gaussian_kde, zscore
    SCIPY_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency
    griddata = None
    ndimage = None
    gaussian_kde = None
    RBFInterpolator = None
    CubicSpline = None
    RegularGridInterpolator = None
    zscore = None
    SCIPY_AVAILABLE = False

# Optional sklearn for advanced interpolation and validation
try:
    from sklearn.gaussian_process import GaussianProcessRegressor
    from sklearn.gaussian_process.kernels import RBF as RBFKernel, Matern, WhiteKernel
    from sklearn.model_selection import cross_val_score, KFold
    SKLEARN_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency
    GaussianProcessRegressor = None
    RBFKernel = None
    Matern = None
    WhiteKernel = None
    cross_val_score = None
    KFold = None
    SKLEARN_AVAILABLE = False

# Logging setup -------------------------------------------------------------
LOGGER_NAME = "custom_map"
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(LOGGER_NAME)
logger.setLevel(logging.INFO)

# Optional MDF support ------------------------------------------------------
try:
    from asammdf import MDF
    try:
        from asammdf.blocks.utils import MultipleOccurrences, MdfException
    except Exception:  # pragma: no cover - different asammdf versions
        try:
            from asammdf.generic import MultipleOccurrences  # type: ignore

            class MdfException(Exception):
                pass
        except Exception:
            class MultipleOccurrences(Exception):
                pass

            class MdfException(Exception):
                pass
    try:
        from asammdf import __version__ as ASAMMDF_VERSION
    except Exception:
        ASAMMDF_VERSION = "N/A"
except Exception:  # pragma: no cover - optional dependency
    MDF = None  # type: ignore
    MultipleOccurrences = None  # type: ignore
    MdfException = None  # type: ignore
    ASAMMDF_VERSION = "N/A"
    logger.warning("asammdf not available. MDF support disabled.")

# Constants -----------------------------------------------------------------
DEFAULT_RPM_BINS = np.arange(0, 8001, 250)
DEFAULT_TQ_BINS = np.arange(-200, 2001, 10)
MIN_SAMPLES_PER_BIN = 6
MAX_FILE_SIZE_MB = 500
CHUNK_SIZE = 100_000

# Advanced filtering constants (MATLAB-level quality)
STEADY_STATE_RPM_TOLERANCE = 50  # RPM variation tolerance for steady-state
STEADY_STATE_TORQUE_TOLERANCE = 10  # Torque variation tolerance (%)
STEADY_STATE_MIN_DURATION = 2.0  # Minimum seconds for steady-state
OUTLIER_Z_THRESHOLD = 3.5  # Modified Z-score threshold
OUTLIER_IQR_MULTIPLIER = 1.5  # IQR multiplier for outlier detection

# Import centralized signal mapping system
try:
    from signal_mapping import SIGNAL_MAP, find_signal_advanced, find_signal_by_role
    
    # Map signal roles to our internal role names
    REQUIRED_SIGNALS: Dict[str, List[str]] = {
        "rpm": SIGNAL_MAP.get("rpm", []),
        "torque": SIGNAL_MAP.get("torque", []),
        "fuel_vol_consumption": SIGNAL_MAP.get("fuel_rate", []),
        "air_mass_flow": SIGNAL_MAP.get("air_mass_flow", []),
        "lambda_raw": SIGNAL_MAP.get("lambda", []),
        "exhaust_temp": SIGNAL_MAP.get("exhaust_temp", []),
        "map_sensor": SIGNAL_MAP.get("map_sensor", []),
        "intake_air_temp_c": SIGNAL_MAP.get("intake_air_temp", []),
        "coolant_temp": SIGNAL_MAP.get("coolant_temp", []),
        "oil_temp": SIGNAL_MAP.get("oil_temp", []),
        "batt_voltage": SIGNAL_MAP.get("battery_voltage", []),
    }
    
    # Ensure we have fallback candidates if signal_mapping doesn't have all
    if not REQUIRED_SIGNALS.get("rpm"):
        REQUIRED_SIGNALS["rpm"] = ["Epm_nEng", "Epm_nEng_RTE", "EngineSpeed", "rpm", "nEng"]
    if not REQUIRED_SIGNALS.get("torque"):
        REQUIRED_SIGNALS["torque"] = ["TqSys_tqCkEngReal", "EngineTorque", "Torque", "Tq"]
    
except ImportError:
    # Fallback if signal_mapping not available
    logger.warning("signal_mapping module not found, using basic signal lists")
    REQUIRED_SIGNALS: Dict[str, List[str]] = {
        "rpm": [
            "Epm_nEng", "Epm_nEng_RTE", "Ext_nEng_RTE", "EngSpeed", "rpm", "engine_rpm", "n_engine",
            "inRpm", "outRpmSpeed", "inRpmSpeed2", "nEng", "EngineSpeed", "n_Eng", "Engine_RPM"
        ],
        "torque": [
            "TqSys_tqCkEngReal", "TqSys_tqCkEngReal_RTE", "EngineTorque", "torque", "Torque",
            "inTorque", "tqEng", "trqEng", "Trq", "Tq", "Engine_Torque"
        ],
        "fuel_vol_consumption": ["FuelRate", "fuel_flow", "FuelCons", "FuCns_volFuCnsTot", "FuelRate"],
        "air_mass_flow": ["air_mass_flow", "MAF_gps", "InM_mfAirCanPurgEstim", "mAir", "AirFlow"],
        "lambda_raw": ["afr", "lambda", "AirFuelRatio", "Lambda", "AFR", "airFuelRatio"],
        "exhaust_temp": ["ExM_tExMnEstim_RTE", "exhaust_temp", "TExh", "EGT", "ExhTemp", "ExhaustTemp"],
        "map_sensor": ["boost", "MAP_kPa", "map", "Boost_kPa", "MAP", "ManifoldPressure"],
        "intake_air_temp_c": ["IAT_C", "intake_air_temp", "Temp_AirInlet", "IAT", "IntakeTemp"],
        "coolant_temp": ["ECT_C", "CoolantTemp", "temp_coolant", "ECT", "EngineTemp"],
        "oil_temp": ["OilTemp", "temp_oil", "Oil_Temp"],
        "batt_voltage": ["battery_voltage", "VBatt", "u_batt", "BatteryVoltage", "VBat"],
    }
    find_signal_by_role = None

CRITICAL_SIGNALS = {"rpm", "torque"}

PRESET_TEMPLATES: Dict[str, Dict[str, Any]] = {
    "ci_engine_default": {
        "label": "CI Engine — BSFC / Emissions (MATLAB Reference)",
        "x_role": "rpm",
        "y_role": "torque",
        "z_roles": ["bsfc_gpkwh", "fuel_mass_flow_kgps", "thermal_efficiency", "bmep_kpa"],
        "x_bins": "800:4500:100",  # Typical CI engine RPM range
        "y_bins": "0:800:10",  # Higher torque range for CI engines
        "min_samples_per_bin": 8,
        "interp_method": "cubic",  # Better for smooth CI engine maps
        "smoothing": 0.8,
        "filter_steady_state": True,  # Critical for CI engine calibration
        "filter_outliers": True,
        "maps": ["engine_bsfc", "exhaust_temperature", "air_fuel_ratio", "thermal_efficiency", "bmep"],
    },
    "ci_engine_advanced": {
        "label": "CI Engine — Advanced (Kriging with Uncertainty)",
        "x_role": "rpm",
        "y_role": "torque",
        "z_roles": ["bsfc_gpkwh", "thermal_efficiency"],
        "x_bins": "800:4500:50",  # Higher resolution
        "y_bins": "0:800:5",
        "min_samples_per_bin": 10,
        "interp_method": "kriging",  # Advanced interpolation with uncertainty
        "smoothing": 0.6,
        "filter_steady_state": True,
        "filter_outliers": True,
        "maps": ["engine_bsfc", "thermal_efficiency"],
    },
    "si_engine_default": {
        "label": "SI Engine — Efficiency / AFR (MATLAB Reference)",
        "x_role": "rpm",
        "y_role": "torque",
        "z_roles": ["mech_power_kw", "bsfc_gpkwh", "lambda_raw", "thermal_efficiency", "volumetric_efficiency"],
        "x_bins": "500:7000:100",  # Wider RPM range for SI engines
        "y_bins": "0:300:5",  # Lower torque range
        "min_samples_per_bin": 6,
        "interp_method": "cubic",
        "smoothing": 0.6,
        "filter_steady_state": True,  # Important for SI calibration
        "filter_outliers": True,
        "maps": ["engine_bsfc", "air_fuel_ratio", "thermal_efficiency", "volumetric_efficiency"],
    },
    "si_engine_advanced": {
        "label": "SI Engine — Advanced (Kriging with Uncertainty)",
        "x_role": "rpm",
        "y_role": "torque",
        "z_roles": ["bsfc_gpkwh", "thermal_efficiency", "volumetric_efficiency"],
        "x_bins": "500:7000:50",
        "y_bins": "0:300:3",
        "min_samples_per_bin": 8,
        "interp_method": "kriging",
        "smoothing": 0.5,
        "filter_steady_state": True,
        "filter_outliers": True,
        "maps": ["engine_bsfc", "thermal_efficiency", "volumetric_efficiency"],
    },
    "electric_motor_default": {
        "label": "Electric Motor — Efficiency",
        "x_role": "rpm",
        "y_role": "torque",
        "z_roles": ["efficiency", "elec_power_kw"],
        "x_bins": "0:20000:200",
        "y_bins": "-400:400:10",
        "min_samples_per_bin": 6,
        "interp_method": "linear",
        "smoothing": 1.0,
        "filter_steady_state": False,  # Less critical for motors
        "filter_outliers": True,
        "maps": ["motor_efficiency"],
    },
    "afr_wide": {
        "label": "AFR Wide",
        "x_role": "rpm",
        "y_role": "torque",
        "z_roles": ["lambda_raw", "air_fuel_ratio"],
        "x_bins": "500:7000:100",
        "y_bins": "0:1:0.02",
        "min_samples_per_bin": 10,
        "interp_method": "linear",
        "smoothing": 0.5,
        "filter_steady_state": True,
        "filter_outliers": True,
        "maps": ["air_fuel_ratio"],
    },
    "emissions_map": {
        "label": "Emissions Map (NOx, PM)",
        "x_role": "rpm",
        "y_role": "torque",
        "z_roles": ["exhaust_temp", "lambda_raw"],
        "x_bins": "800:4500:100",
        "y_bins": "0:600:10",
        "min_samples_per_bin": 8,
        "interp_method": "cubic",
        "smoothing": 0.7,
        "filter_steady_state": True,
        "filter_outliers": True,
        "maps": ["exhaust_temperature", "air_fuel_ratio"],
    },
}

# Utility helpers -----------------------------------------------------------
def find_signal(columns: List[str], role: str, overrides: Optional[Dict[str, str]] = None, mdf=None) -> Optional[str]:
    """
    Find signal by role using advanced signal mapping system.
    
    This function now uses the centralized signal_mapping module for better detection.
    """
    cols = set(columns)
    if overrides and role in overrides and overrides[role] in cols:
        return overrides[role]

    # Use centralized signal mapping if available (preferred method)
    try:
        if find_signal_by_role and mdf is not None:
            # Map internal role names to signal_mapping role names
            role_mapping = {
                "rpm": "rpm",
                "torque": "torque",
                "lambda_raw": "lambda",
                "intake_air_temp_c": "intake_air_temp",
                "coolant_temp": "coolant_temp",
                "exhaust_temp": "exhaust_temp",
                "map_sensor": "map_sensor",
                "fuel_vol_consumption": "fuel_rate",
                "air_mass_flow": "air_mass_flow",
                "oil_temp": "oil_temp",
                "batt_voltage": "battery_voltage",
            }
            mapped_role = role_mapping.get(role, role)
            result = find_signal_by_role(mdf, mapped_role)
            if result:
                return result
    except Exception:
        pass  # Fall back to advanced matching
    
    # Use advanced signal mapping if available (without MDF)
    try:
        if find_signal_advanced:
            # Map role to signal_mapping role names
            role_mapping = {
                "rpm": "rpm",
                "torque": "torque",
                "lambda_raw": "lambda",
                "intake_air_temp_c": "intake_air_temp",
                "coolant_temp": "coolant_temp",
                "exhaust_temp": "exhaust_temp",
                "map_sensor": "map_sensor",
                "fuel_vol_consumption": "fuel_rate",
                "air_mass_flow": "air_mass_flow",
                "oil_temp": "oil_temp",
                "batt_voltage": "battery_voltage",
            }
            mapped_role = role_mapping.get(role, role)
            result = find_signal_advanced(columns, mapped_role, fuzzy_match=True, substring_match=True)
            if result:
                return result
    except Exception:
        pass  # Fall back to basic matching

    candidates = REQUIRED_SIGNALS.get(role, [])
    # First try exact match
    for candidate in candidates:
        if candidate in cols:
            return candidate

    # Then try case-insensitive exact match
    lower_map = {c.lower(): c for c in columns}
    for candidate in candidates:
        cand_lower = candidate.lower()
        if cand_lower in lower_map:
            return lower_map[cand_lower]
    
    # Then try substring match (candidate in column name)
    for candidate in candidates:
        cand_lower = candidate.lower()
        for col_lower, original in lower_map.items():
            if cand_lower in col_lower or col_lower in cand_lower:
                return original
    
    # Last resort: fuzzy match on key parts (e.g., "nEng" in "Epm_nEng")
    key_parts = {
        "rpm": ["rpm", "neng", "engspeed", "speed"],
        "torque": ["torque", "tq", "trq", "moment"],
    }
    if role in key_parts:
        for key_part in key_parts[role]:
            for col_lower, original in lower_map.items():
                if key_part in col_lower:
                    return original
    
    return None


class MapGeneratorError(Exception):
    pass


def detect_outliers_advanced(
    df: pd.DataFrame, 
    columns: List[str], 
    method: str = "combined",
    z_threshold: float = OUTLIER_Z_THRESHOLD,
    iqr_multiplier: float = OUTLIER_IQR_MULTIPLIER
) -> pd.Series:
    """
    Advanced outlier detection using multiple methods (MATLAB-level robustness).
    
    Methods:
    - 'zscore': Modified Z-score (more robust to outliers)
    - 'iqr': Interquartile Range method
    - 'combined': Uses both methods (OR logic)
    """
    if df.empty or not columns:
        return pd.Series(True, index=df.index)
    
    valid_mask = pd.Series(True, index=df.index)
    
    for col in columns:
        if col not in df.columns:
            continue
        series = pd.to_numeric(df[col], errors="coerce")
        if series.notna().sum() < 10:  # Need minimum data
            continue
        
        # Method 1: Modified Z-score (more robust)
        if method in ("zscore", "combined"):
            median = series.median()
            mad = (series - median).abs().median()  # Median Absolute Deviation
            if mad > 0:
                modified_z = 0.6745 * (series - median) / mad  # 0.6745 makes it comparable to std
                z_outliers = np.abs(modified_z) > z_threshold
            else:
                z_outliers = pd.Series(False, index=series.index)
        
        # Method 2: IQR method
        if method in ("iqr", "combined"):
            q1 = series.quantile(0.25)
            q3 = series.quantile(0.75)
            iqr = q3 - q1
            if iqr > 0:
                lower_bound = q1 - iqr_multiplier * iqr
                upper_bound = q3 + iqr_multiplier * iqr
                iqr_outliers = (series < lower_bound) | (series > upper_bound)
            else:
                iqr_outliers = pd.Series(False, index=series.index)
        
        # Combine methods
        if method == "combined":
            col_outliers = z_outliers | iqr_outliers
        elif method == "zscore":
            col_outliers = z_outliers
        else:  # iqr
            col_outliers = iqr_outliers
        
        valid_mask = valid_mask & ~col_outliers
    
    return valid_mask


def detect_steady_state_regions(
    df: pd.DataFrame,
    rpm_col: str = "rpm",
    torque_col: str = "torque",
    time_col: Optional[str] = None,
    rpm_tolerance: float = STEADY_STATE_RPM_TOLERANCE,
    torque_tolerance_pct: float = STEADY_STATE_TORQUE_TOLERANCE,
    min_duration: float = STEADY_STATE_MIN_DURATION
) -> pd.Series:
    """
    Detect steady-state operating regions (critical for accurate map generation).
    Based on MATLAB CI/SI engine calibration best practices.
    """
    if rpm_col not in df.columns or torque_col not in df.columns:
        return pd.Series(True, index=df.index)
    
    # Create time index if not available
    if time_col and time_col in df.columns:
        time_series = pd.to_numeric(df[time_col], errors="coerce")
    else:
        time_series = pd.Series(range(len(df)), index=df.index)
        if time_series.diff().mean() == 0:
            # Estimate time from index spacing (assuming constant sampling)
            time_series = time_series * 0.01  # Assume ~100 Hz sampling
    
    rpm = pd.to_numeric(df[rpm_col], errors="coerce")
    torque = pd.to_numeric(df[torque_col], errors="coerce")
    
    steady_mask = pd.Series(False, index=df.index)
    
    # Calculate rolling statistics
    window_size = max(10, int(len(df) * 0.02))  # ~2% of data or min 10
    
    rpm_rolling_std = rpm.rolling(window=window_size, center=True).std()
    rpm_rolling_mean = rpm.rolling(window=window_size, center=True).mean()
    
    torque_rolling_mean = torque.rolling(window=window_size, center=True).mean()
    torque_rolling_std = torque.rolling(window=window_size, center=True).std()
    
    # Steady-state conditions:
    # 1. RPM variation within tolerance
    rpm_steady = rpm_rolling_std < rpm_tolerance
    
    # 2. Torque variation within percentage tolerance
    torque_steady = (torque_rolling_std / (torque_rolling_mean.abs() + 1e-6) * 100) < torque_tolerance_pct
    
    # 3. Both RPM and torque must be steady
    steady_mask = rpm_steady & torque_steady
    
    # Filter by minimum duration (must maintain steady-state for min_duration)
    if time_series.diff().sum() > 0:
        # Group consecutive steady-state points
        steady_groups = (steady_mask != steady_mask.shift()).cumsum()
        for group_id in steady_groups.unique():
            group_mask = steady_groups == group_id
            if steady_mask[group_mask].any():
                group_duration = time_series[group_mask].max() - time_series[group_mask].min()
                if group_duration < min_duration:
                    steady_mask.loc[group_mask] = False
    
    return steady_mask


def validate_data_quality(df: pd.DataFrame, required: Optional[List[str]] = None) -> Dict[str, Any]:
    required = required or ["rpm", "torque"]
    report: Dict[str, Any] = {
        "missing_signals": [],
        "signal_stats": {},
        "total_samples": int(len(df)),
        "valid_samples": {},
        "data_quality_issues": [],
        "outlier_info": {},
        "steady_state_info": {},
    }
    
    for signal in required:
        if signal not in df.columns:
            report["missing_signals"].append(signal)
            continue
        series = pd.to_numeric(df[signal], errors="coerce")
        valid = series.notna().sum()
        report["valid_samples"][signal] = int(valid)
        if valid == 0:
            report["data_quality_issues"].append(f"Signal '{signal}' has no valid samples")
            continue
        
        # Advanced statistics
        stats = {
            "min": float(series.min()),
            "max": float(series.max()),
            "mean": float(series.mean()),
            "std": float(series.std()),
            "median": float(series.median()),
            "q25": float(series.quantile(0.25)),
            "q75": float(series.quantile(0.75)),
            "iqr": float(series.quantile(0.75) - series.quantile(0.25)),
            "skewness": float(series.skew()) if valid > 10 else None,
            "kurtosis": float(series.kurtosis()) if valid > 10 else None,
            "valid_percentage": float(valid / len(series) * 100.0),
        }
        report["signal_stats"][signal] = stats
        
        # Outlier detection
        if SCIPY_AVAILABLE and zscore is not None:
            try:
                outliers = detect_outliers_advanced(df, [signal], method="combined")
                outlier_count = (~outliers).sum()
                report["outlier_info"][signal] = {
                    "outlier_count": int(outlier_count),
                    "outlier_percentage": float(outlier_count / len(df) * 100.0),
                }
            except Exception:
                pass
    
    # Steady-state analysis (if RPM and torque available)
    if "rpm" in df.columns and "torque" in df.columns:
        try:
            steady_mask = detect_steady_state_regions(df)
            steady_count = steady_mask.sum()
            report["steady_state_info"] = {
                "steady_state_samples": int(steady_count),
                "steady_state_percentage": float(steady_count / len(df) * 100.0),
                "transient_samples": int(len(df) - steady_count),
            }
        except Exception:
            pass
    
    return report


def derive_signals(df: pd.DataFrame, overrides: Optional[Dict[str, str]] = None) -> Tuple[pd.DataFrame, Dict[str, str], List[Dict[str, Any]], Dict[str, List[str]]]:
    df = df.copy()
    mapping: Dict[str, str] = {}
    report_rows: List[Dict[str, Any]] = []
    missing: Dict[str, List[str]] = {}

    for role in REQUIRED_SIGNALS:
        matched = find_signal(list(df.columns), role, overrides)
        if matched:
            mapping[role] = matched
            df[role] = pd.to_numeric(df[matched], errors="coerce")
            report_rows.append({"standard_signal": role, "csv_column": matched, "present": True})
        else:
            missing[role] = REQUIRED_SIGNALS[role]
            report_rows.append({"standard_signal": role, "csv_column": "Not Found", "present": False})

    if "rpm" in df.columns:
        df["omega_rad_s"] = df["rpm"] * (2.0 * math.pi / 60.0)
    else:
        df["omega_rad_s"] = np.nan

    if {"torque", "omega_rad_s"}.issubset(df.columns):
        df["mech_power_kw"] = df["torque"] * df["omega_rad_s"] / 1000.0
    else:
        df["mech_power_kw"] = np.nan

    df["fuel_mass_flow_kgps"] = np.nan
    if "fuel_vol_consumption" in df.columns:
        series = df["fuel_vol_consumption"].abs()
        median_val = series.median(skipna=True)
        if 0 < median_val < 200:  # assume L/h
            density = 0.745
            df["fuel_mass_flow_kgps"] = df["fuel_vol_consumption"] * density / 3600.0
        elif median_val >= 200:
            density = 0.745
            df["fuel_mass_flow_kgps"] = df["fuel_vol_consumption"] * density / 3_600_000.0

    df["bsfc_gpkwh"] = np.nan
    valid_mask = (
        df["mech_power_kw"].notna()
        & (df["mech_power_kw"] > 0.1)
        & df["fuel_mass_flow_kgps"].notna()
    )
    if valid_mask.any():
        df.loc[valid_mask, "bsfc_gpkwh"] = (
            df.loc[valid_mask, "fuel_mass_flow_kgps"] * 3_600_000.0 / df.loc[valid_mask, "mech_power_kw"]
        )

    # Enhanced volumetric efficiency calculation (MATLAB-level accuracy)
    if {"map_sensor", "intake_air_temp_c", "air_mass_flow", "rpm"}.issubset(df.columns):
        R = 287.05  # Gas constant J/(kg·K)
        mask = df[["map_sensor", "intake_air_temp_c", "air_mass_flow", "rpm"]].notna().all(axis=1)
        df["volumetric_efficiency"] = np.nan
        if mask.any():
            pressure_pa = df.loc[mask, "map_sensor"] * 1000.0  # Convert kPa to Pa
            temp_k = df.loc[mask, "intake_air_temp_c"] + 273.15
            displacement_l = 2.0  # Assumed displacement, could be made configurable
            # Theoretical air flow for 4-stroke engine: m_dot = (P * V_d * N) / (2 * 60 * R * T)
            # P in Pa, V_d in m³, N in RPM, result in kg/s
            displacement_m3 = displacement_l * 0.001
            theoretical = (pressure_pa * displacement_m3 * df.loc[mask, "rpm"]) / (2 * 60 * R * temp_k)
            df.loc[mask, "volumetric_efficiency"] = (
                df.loc[mask, "air_mass_flow"] / (theoretical + 1e-6)
            ).clip(lower=0.1, upper=2.5)  # Reasonable bounds for volumetric efficiency
        else:
            df["volumetric_efficiency"] = np.nan
    
    # Enhanced thermal efficiency calculation (MATLAB-level)
    df["thermal_efficiency"] = np.nan
    if {"mech_power_kw", "fuel_mass_flow_kgps"}.issubset(df.columns):
        mask = df[["mech_power_kw", "fuel_mass_flow_kgps"]].notna().all(axis=1) & (df["mech_power_kw"] > 0.1)
        if mask.any():
            # Thermal efficiency = mechanical power / (fuel mass flow * LHV)
            # LHV (Lower Heating Value) for typical fuels: diesel ~42.5 MJ/kg, gasoline ~44 MJ/kg
            lhv_diesel = 42.5e6  # J/kg
            lhv_gasoline = 44e6  # J/kg
            # Use diesel LHV as default (can be made configurable based on fuel type)
            fuel_power_mw = df.loc[mask, "fuel_mass_flow_kgps"] * lhv_diesel / 1e6  # MW
            df.loc[mask, "thermal_efficiency"] = (
                df.loc[mask, "mech_power_kw"] / (fuel_power_mw * 1000 + 1e-6)
            ).clip(lower=0.0, upper=0.6)  # Reasonable bounds for thermal efficiency (0-60%)
    
    # BMEP (Brake Mean Effective Pressure) calculation (MATLAB-level)
    df["bmep_kpa"] = np.nan
    if {"torque", "rpm"}.issubset(df.columns):
        mask = df[["torque", "rpm"]].notna().all(axis=1) & (df["rpm"] > 100) & (df["torque"] > 0)
        if mask.any():
            displacement_m3 = 2.0 * 0.001  # Convert L to m³
            # BMEP = (2π * T * n_r) / (V_d * n_r) = 2π * T / V_d
            # Where T is torque in N·m, V_d is displacement in m³
            # Result in Pa, convert to kPa
            df.loc[mask, "bmep_kpa"] = (
                2 * math.pi * df.loc[mask, "torque"] / (displacement_m3 + 1e-6) / 1000.0
            ).clip(lower=0.0, upper=2000.0)  # Reasonable bounds for BMEP (0-2000 kPa)
    
    # Mean piston speed calculation (MATLAB-level)
    df["mean_piston_speed_ms"] = np.nan
    if "rpm" in df.columns:
        mask = df["rpm"].notna() & (df["rpm"] > 0)
        if mask.any():
            stroke_m = 0.08  # Assumed stroke length (could be made configurable)
            # Mean piston speed = 2 * stroke * RPM / 60
            df.loc[mask, "mean_piston_speed_ms"] = (
                2 * stroke_m * df.loc[mask, "rpm"] / 60.0
            ).clip(lower=0.0, upper=25.0)  # Reasonable bounds (0-25 m/s)

    df["efficiency"] = df.get("efficiency", np.nan)
    if "batt_voltage" not in df.columns:
        df["batt_voltage"] = np.nan

    return df, mapping, report_rows, missing


def apply_smoothing(grid: np.ndarray, sigma: float, valid_mask: np.ndarray) -> np.ndarray:
    if sigma <= 0 or ndimage is None:
        return grid
    smoothed = grid.copy()
    smoothed[~valid_mask] = 0
    smoothed = ndimage.gaussian_filter(smoothed, sigma=sigma)
    weights = np.ones_like(grid)
    weights[~valid_mask] = 0
    weights = ndimage.gaussian_filter(weights, sigma=sigma)
    with np.errstate(invalid="ignore"):
        smoothed = np.divide(smoothed, weights, where=weights > 1e-6)
    smoothed[~valid_mask] = np.nan
    return smoothed


def create_calibration_map(
    df: pd.DataFrame,
    x_col: str,
    y_col: str,
    value_col: str,
    x_bins: Optional[np.ndarray] = None,
    y_bins: Optional[np.ndarray] = None,
    min_samples_per_bin: int = MIN_SAMPLES_PER_BIN,
    interp_method: str = "linear",
    grid_resolution: Tuple[int, int] = (100, 100),
    smoothing: float = 0.5,
    contour_levels: int = 10,
    enable_contours: bool = True,
    enable_surface: bool = True,
    filter_steady_state: bool = False,
    filter_outliers: bool = False,
) -> Dict[str, Any]:
    if not {x_col, y_col, value_col}.issubset(df.columns):
        raise ValueError(f"DataFrame must contain {x_col}, {y_col}, {value_col}")

    # Advanced data preprocessing (MATLAB-level quality filtering)
    df_clean = df[[x_col, y_col, value_col]].apply(pd.to_numeric, errors="coerce").dropna()
    
    if df_clean.empty:
        raise ValueError("No valid data points after dropping NaNs.")
    
    original_count = len(df_clean)
    
    # Apply steady-state filtering if requested (critical for CI/SI engine calibration)
    if filter_steady_state and x_col == "rpm" and y_col == "torque":
        try:
            steady_mask = detect_steady_state_regions(df_clean, rpm_col=x_col, torque_col=y_col)
            df_clean = df_clean[steady_mask].copy()
            steady_count = len(df_clean)
            logger.info(f"Steady-state filtering: {original_count} → {steady_count} samples ({steady_count/original_count*100:.1f}%)")
        except Exception as e:
            logger.warning(f"Steady-state filtering failed: {e}")
    
    # Apply outlier filtering if requested
    if filter_outliers:
        try:
            outlier_mask = detect_outliers_advanced(df_clean, [x_col, y_col, value_col], method="combined")
            df_clean = df_clean[outlier_mask].copy()
            filtered_count = len(df_clean)
            logger.info(f"Outlier filtering: {original_count} → {filtered_count} samples ({filtered_count/original_count*100:.1f}%)")
        except Exception as e:
            logger.warning(f"Outlier filtering failed: {e}")
    
    if df_clean.empty:
        raise ValueError("No valid data points after filtering.")

    # Ensure bins are sorted and unique
    if x_bins is not None:
        x_edges = np.unique(np.sort(x_bins))
        if len(x_edges) < 2:
            x_edges = np.linspace(df_clean[x_col].min(), df_clean[x_col].max(), 21)
    else:
        x_edges = np.linspace(df_clean[x_col].min(), df_clean[x_col].max(), 21)
    
    if y_bins is not None:
        y_edges = np.unique(np.sort(y_bins))
        if len(y_edges) < 2:
            y_edges = np.linspace(df_clean[y_col].min(), df_clean[y_col].max(), 21)
    else:
        y_edges = np.linspace(df_clean[y_col].min(), df_clean[y_col].max(), 21)

    df_clean["x_bin"] = pd.cut(df_clean[x_col], bins=x_edges, labels=False, include_lowest=True)
    df_clean["y_bin"] = pd.cut(df_clean[y_col], bins=y_edges, labels=False, include_lowest=True)

    # Advanced aggregation with percentiles and quartiles (MATLAB-level statistics)
    def percentile_25(x):
        return x.quantile(0.25)
    def percentile_75(x):
        return x.quantile(0.75)
    def percentile_10(x):
        return x.quantile(0.10)
    def percentile_90(x):
        return x.quantile(0.90)
    def percentile_5(x):
        return x.quantile(0.05)
    def percentile_95(x):
        return x.quantile(0.95)
    
    agg = (
        df_clean.groupby(["x_bin", "y_bin"], observed=False)[value_col]
        .agg([
            "count", "mean", "median", "std", "min", "max",
            percentile_25, percentile_75, percentile_10, percentile_90,
            percentile_5, percentile_95
        ])
        .reset_index()
    )
    
    # Rename columns for clarity
    agg.columns = [
        "x_bin", "y_bin", "count", "mean", "median", "std", "min", "max",
        "p25", "p75", "p10", "p90", "p5", "p95"
    ]

    x_centers = 0.5 * (x_edges[:-1] + x_edges[1:])
    y_centers = 0.5 * (y_edges[:-1] + y_edges[1:])
    grid_shape = (len(y_centers), len(x_centers))
    grid_mean = np.full(grid_shape, np.nan)
    grid_median = np.full_like(grid_mean, np.nan)
    grid_std = np.full_like(grid_mean, np.nan)
    grid_min = np.full_like(grid_mean, np.nan)
    grid_max = np.full_like(grid_mean, np.nan)
    grid_p25 = np.full_like(grid_mean, np.nan)
    grid_p75 = np.full_like(grid_mean, np.nan)
    grid_p10 = np.full_like(grid_mean, np.nan)
    grid_p90 = np.full_like(grid_mean, np.nan)
    grid_p5 = np.full_like(grid_mean, np.nan)
    grid_p95 = np.full_like(grid_mean, np.nan)
    grid_count = np.zeros_like(grid_mean, dtype=int)
    grid_iqr = np.full_like(grid_mean, np.nan)  # Interquartile range

    for _, row in agg.iterrows():
        ix = int(row["x_bin"])
        iy = int(row["y_bin"])
        if 0 <= ix < len(x_centers) and 0 <= iy < len(y_centers):
            count = int(row["count"])
            grid_count[iy, ix] = count
            if count >= min_samples_per_bin:
                grid_mean[iy, ix] = row["mean"]
                grid_median[iy, ix] = row["median"]
                grid_std[iy, ix] = row["std"] if pd.notna(row["std"]) else np.nan
                grid_min[iy, ix] = row["min"]
                grid_max[iy, ix] = row["max"]
                grid_p25[iy, ix] = row["p25"]
                grid_p75[iy, ix] = row["p75"]
                grid_p10[iy, ix] = row["p10"]
                grid_p90[iy, ix] = row["p90"]
                grid_p5[iy, ix] = row["p5"]
                grid_p95[iy, ix] = row["p95"]
                # Calculate IQR
                if pd.notna(row["p75"]) and pd.notna(row["p25"]):
                    grid_iqr[iy, ix] = row["p75"] - row["p25"]

    valid_mask = grid_count >= min_samples_per_bin
    mean_map = np.where(valid_mask, grid_mean, np.nan)
    median_map = np.where(valid_mask, grid_median, np.nan)

    if smoothing > 0:
        mean_map = apply_smoothing(mean_map, smoothing, valid_mask)
        median_map = apply_smoothing(median_map, smoothing, valid_mask)

    contour_data = None
    if enable_contours and np.sum(~np.isnan(mean_map)) > 10:
        X, Y = np.meshgrid(x_centers, y_centers)
        contour_data = {"x": X, "y": Y, "z": mean_map, "levels": contour_levels}

    # Always generate surface_data - it's required for surface plots
    # Generate surface data - can use simple meshgrid if scipy not available
    surface_data = None
    # Create meshgrid once for reuse (for fallback case)
    X_mesh, Y_mesh = np.meshgrid(x_centers, y_centers)
    
    if enable_surface:
        if griddata is not None and gaussian_kde is not None:
            # Use scipy interpolation for smoother surface
            points: List[List[float]] = []
            values: List[float] = []
            for iy, y_center in enumerate(y_centers):
                for ix, x_center in enumerate(x_centers):
                    if valid_mask[iy, ix] and not np.isnan(grid_mean[iy, ix]):
                        points.append([x_center, y_center])
                        values.append(grid_mean[iy, ix])
            
            if len(points) >= 10:
                pts = np.array(points)
                vals = np.array(values)
                grid_x, grid_y = np.meshgrid(
                    np.linspace(pts[:, 0].min(), pts[:, 0].max(), grid_resolution[0]),
                    np.linspace(pts[:, 1].min(), pts[:, 1].max(), grid_resolution[1]),
                )
                try:
                    # Advanced interpolation methods (MATLAB-level with uncertainty quantification)
                    if interp_method == "kriging" and SKLEARN_AVAILABLE:
                        # Gaussian Process Regression (Kriging) with uncertainty
                        try:
                            # Use RBF kernel with white noise kernel for uncertainty
                            kernel = RBFKernel(length_scale=1.0) + WhiteKernel(noise_level=0.1)
                            gpr = GaussianProcessRegressor(kernel=kernel, alpha=0.01, n_restarts_optimizer=3)
                            gpr.fit(pts, vals)
                            grid_xy = np.column_stack([grid_x.ravel(), grid_y.ravel()])
                            grid_z, grid_z_std = gpr.predict(grid_xy, return_std=True)
                            grid_z = grid_z.reshape(grid_x.shape)
                            grid_z_std = grid_z_std.reshape(grid_x.shape)
                            # Store uncertainty (95% confidence interval)
                            grid_z_lower = grid_z - 1.96 * grid_z_std
                            grid_z_upper = grid_z + 1.96 * grid_z_std
                            surface_data = {
                                "x": grid_x, "y": grid_y, "z": grid_z,
                                "z_std": grid_z_std,
                                "z_lower_ci": grid_z_lower,  # Lower 95% CI
                                "z_upper_ci": grid_z_upper,  # Upper 95% CI
                            }
                            logger.info("Kriging interpolation with uncertainty quantification complete")
                        except Exception as e:
                            logger.warning("Kriging interpolation failed, falling back to RBF: %s", e)
                            interp_method = "rbf"  # Fall through to RBF
                    
                    if interp_method == "rbf" and SCIPY_AVAILABLE and RBFInterpolator is not None:
                        try:
                            rbf = RBFInterpolator(pts, vals, kernel='thin_plate_spline', smoothing=0.1)
                            grid_xy = np.column_stack([grid_x.ravel(), grid_y.ravel()])
                            grid_z = rbf(grid_xy).reshape(grid_x.shape)
                            surface_data = {"x": grid_x, "y": grid_y, "z": grid_z}
                        except Exception as e:
                            logger.warning("RBF interpolation failed, falling back to cubic: %s", e)
                            grid_z = griddata(pts, vals, (grid_x, grid_y), method="cubic")
                            surface_data = {"x": grid_x, "y": grid_y, "z": grid_z}
                    elif interp_method == "cubic_spline" and SCIPY_AVAILABLE and CubicSpline is not None:
                        # Use cubic spline interpolation
                        grid_z = griddata(pts, vals, (grid_x, grid_y), method="cubic")
                        surface_data = {"x": grid_x, "y": grid_y, "z": grid_z}
                    else:
                        # Standard griddata methods: linear, cubic, nearest
                        grid_z = griddata(pts, vals, (grid_x, grid_y), method=interp_method)
                        if surface_data is None or "z_std" not in surface_data:
                            surface_data = {"x": grid_x, "y": grid_y, "z": grid_z}
                    
                    # Apply smoothing if requested (only if not already processed)
                    if smoothing > 0 and ndimage is not None and grid_z is not None and "z_std" not in surface_data:
                        grid_z = ndimage.gaussian_filter(grid_z, sigma=smoothing)
                        if isinstance(surface_data, dict):
                            surface_data["z"] = grid_z
                except Exception as exc:
                    logger.warning("Surface interpolation failed: %s", exc)
                    # Fall through to direct mapping
    # Fallback: use direct mean_map (already on grid) - ALWAYS create surface_data
    if surface_data is None:
        # Use the binned mean_map directly - it's already on a grid
        # Use the meshgrid created above (X_mesh, Y_mesh)
        Z = mean_map.copy()  # This is already aligned with x_centers, y_centers
        
        # Create surface_data - NaN values will be converted to None in create_surface_trace
        surface_data = {"x": X_mesh, "y": Y_mesh, "z": Z}
        
        # Log for debugging
        valid_points = np.sum(~np.isnan(Z))
        total_points = Z.size
        if valid_points == 0:
            logger.warning(f"All surface data values are NaN ({total_points} points) - surface plot may not display correctly")
        else:
            logger.info(f"Created fallback surface data: {valid_points}/{total_points} valid points ({(valid_points/total_points)*100:.1f}%)")
    
    # Ensure surface_data exists even if enable_surface was False
    if not enable_surface and surface_data is None:
        Z = mean_map.copy()
        surface_data = {"x": X_mesh, "y": Y_mesh, "z": Z}
        logger.info("Created surface_data from mean_map (enable_surface was False)")

    # Calculate quality metrics (R², RMSE) for map validation
    quality_metrics = {}
    valid_data = df_clean[[x_col, y_col, value_col]].dropna()
    if len(valid_data) > 10:
        try:
            # Predict values using binned means
            valid_data["x_bin_pred"] = pd.cut(valid_data[x_col], bins=x_edges, labels=False, include_lowest=True)
            valid_data["y_bin_pred"] = pd.cut(valid_data[y_col], bins=y_edges, labels=False, include_lowest=True)
            predictions = []
            for _, row in valid_data.iterrows():
                ix = int(row["x_bin_pred"]) if pd.notna(row["x_bin_pred"]) else None
                iy = int(row["y_bin_pred"]) if pd.notna(row["y_bin_pred"]) else None
                if (ix is not None and iy is not None and 
                    0 <= ix < len(x_centers) and 0 <= iy < len(y_centers) and 
                    valid_mask[iy, ix]):
                    predictions.append(grid_mean[iy, ix])
                else:
                    predictions.append(np.nan)
            valid_data["predicted"] = predictions
            valid_comparison = valid_data[["predicted", value_col]].dropna()
            
            if len(valid_comparison) > 5:
                observed = valid_comparison[value_col].values
                predicted = valid_comparison["predicted"].values
                
                # R² (Coefficient of Determination)
                ss_res = np.sum((observed - predicted) ** 2)
                ss_tot = np.sum((observed - np.mean(observed)) ** 2)
                r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else np.nan
                
                # RMSE (Root Mean Squared Error)
                rmse = np.sqrt(np.mean((observed - predicted) ** 2))
                
                # MAE (Mean Absolute Error)
                mae = np.mean(np.abs(observed - predicted))
                
                # Mean Absolute Percentage Error
                with np.errstate(divide='ignore', invalid='ignore'):
                    mape = np.mean(np.abs((observed - predicted) / observed)) * 100
                
                quality_metrics = {
                    "r_squared": float(r_squared) if np.isfinite(r_squared) else None,
                    "rmse": float(rmse) if np.isfinite(rmse) else None,
                    "mae": float(mae) if np.isfinite(mae) else None,
                    "mape": float(mape) if np.isfinite(mape) else None,
                    "n_samples_validated": int(len(valid_comparison))
                }
        except Exception as e:
            logger.warning("Quality metrics calculation failed: %s", e)
            quality_metrics = {"error": str(e)}
    
    # Enhanced statistics (MATLAB-level) - Convert NaN to None for JSON serialization
    global_mean_val = np.nanmean(mean_map)
    global_std_val = np.nanstd(mean_map)
    global_min_val = np.nanmin(mean_map)
    global_max_val = np.nanmax(mean_map)
    global_median_val = np.nanmedian(mean_map)
    
    stats = {
        "global_mean": float(global_mean_val) if np.isfinite(global_mean_val) else None,
        "global_std": float(global_std_val) if np.isfinite(global_std_val) else None,
        "global_min": float(global_min_val) if np.isfinite(global_min_val) else None,
        "global_max": float(global_max_val) if np.isfinite(global_max_val) else None,
        "global_median": float(global_median_val) if np.isfinite(global_median_val) else None,
        "global_p25": float(np.nanpercentile(mean_map, 25)) if np.any(~np.isnan(mean_map)) else None,
        "global_p75": float(np.nanpercentile(mean_map, 75)) if np.any(~np.isnan(mean_map)) else None,
        "coverage_ratio": float(np.sum(valid_mask) / valid_mask.size) if valid_mask.size else 0.0,
        "quality_metrics": quality_metrics,
    }

    # Generate validation plots (MATLAB-level analysis)
    validation_plots = {}
    try:
        validation_plots_dict = create_validation_plots(
            {}, df_clean, x_col, y_col, value_col,
            x_edges, y_edges, mean_map, valid_mask, x_centers, y_centers
        )
        # Convert Plotly figures to JSON strings for serialization
        for plot_name, fig in validation_plots_dict.items():
            validation_plots[plot_name] = {"plotly_json": fig.to_json()}
    except Exception as e:
        logger.warning("Failed to generate validation plots: %s", e)
    
    return {
        "x_centers": x_centers,
        "y_centers": y_centers,
        "mean_map": mean_map,
        "median_map": median_map,
        "std_map": np.where(valid_mask, grid_std, np.nan),
        "min_map": np.where(valid_mask, grid_min, np.nan),
        "max_map": np.where(valid_mask, grid_max, np.nan),
        "p25_map": np.where(valid_mask, grid_p25, np.nan),
        "p75_map": np.where(valid_mask, grid_p75, np.nan),
        "p10_map": np.where(valid_mask, grid_p10, np.nan),
        "p90_map": np.where(valid_mask, grid_p90, np.nan),
        "p5_map": np.where(valid_mask, grid_p5, np.nan),
        "p95_map": np.where(valid_mask, grid_p95, np.nan),
        "iqr_map": np.where(valid_mask, grid_iqr, np.nan),
        "count_map": grid_count,
        "surface_data": surface_data,
        "contour_data": contour_data,
        "stats": stats,
        "validation_plots": validation_plots,
        "meta": {
            "min_samples_per_bin": int(min_samples_per_bin),
            "cells_with_data": int(np.sum(valid_mask)),
            "total_cells": int(valid_mask.size),
            "smoothing_applied": float(smoothing),
            "interpolation_method": interp_method,
            "total_input_samples": int(len(df_clean)),
            "bins_x": len(x_centers),
            "bins_y": len(y_centers),
        },
    }


def export_map_to_matlab(map_dict: Dict[str, Any], output_path: Path, map_name: str = "map") -> str:
    """
    Export map data to MATLAB .mat format (compatible with scipy.io.savemat).
    This creates a file that can be loaded in MATLAB using: load('filename.mat')
    """
    try:
        from scipy.io import savemat
    except ImportError:
        raise MapGeneratorError("scipy.io.savemat not available for MATLAB export")
    
    x_centers = map_dict.get("x_centers", np.array([]))
    y_centers = map_dict.get("y_centers", np.array([]))
    
    # Prepare data structure for MATLAB
    mat_data = {
        f"{map_name}_x": x_centers,
        f"{map_name}_y": y_centers,
        f"{map_name}_mean": map_dict.get("mean_map", np.array([])),
        f"{map_name}_median": map_dict.get("median_map", np.array([])),
        f"{map_name}_std": map_dict.get("std_map", np.array([])),
        f"{map_name}_count": map_dict.get("count_map", np.array([])),
        f"{map_name}_min": map_dict.get("min_map", np.array([])),
        f"{map_name}_max": map_dict.get("max_map", np.array([])),
        f"{map_name}_p25": map_dict.get("p25_map", np.array([])),
        f"{map_name}_p75": map_dict.get("p75_map", np.array([])),
        f"{map_name}_iqr": map_dict.get("iqr_map", np.array([])),
        f"{map_name}_stats": map_dict.get("stats", {}),
    }
    
    # Add surface data if available
    if map_dict.get("surface_data"):
        surf = map_dict["surface_data"]
        mat_data[f"{map_name}_surface_x"] = surf.get("x", np.array([]))
        mat_data[f"{map_name}_surface_y"] = surf.get("y", np.array([]))
        mat_data[f"{map_name}_surface_z"] = surf.get("z", np.array([]))
    
    savemat(str(output_path), mat_data)
    return str(output_path)


def export_map_to_csv(map_dict: Dict[str, Any], output_path: Path, include_stats: bool = True) -> str:
    """
    Export map data to CSV format with comprehensive statistics per bin.
    Creates a 2D table format suitable for Excel or other tools.
    """
    x_centers = map_dict.get("x_centers", np.array([]))
    y_centers = map_dict.get("y_centers", np.array([]))
    mean_map = map_dict.get("mean_map", np.array([]))
    
    # Create a DataFrame with comprehensive statistics
    rows = []
    for iy, y_val in enumerate(y_centers):
        for ix, x_val in enumerate(x_centers):
            if include_stats:
                row = {
                    "X_Value": float(x_val),
                    "Y_Value": float(y_val),
                    "Mean": float(mean_map[iy, ix]) if not np.isnan(mean_map[iy, ix]) else np.nan,
                    "Median": float(map_dict.get("median_map", np.array([]))[iy, ix]) if not np.isnan(map_dict.get("median_map", np.array([]))[iy, ix]) else np.nan,
                    "StdDev": float(map_dict.get("std_map", np.array([]))[iy, ix]) if not np.isnan(map_dict.get("std_map", np.array([]))[iy, ix]) else np.nan,
                    "Min": float(map_dict.get("min_map", np.array([]))[iy, ix]) if not np.isnan(map_dict.get("min_map", np.array([]))[iy, ix]) else np.nan,
                    "Max": float(map_dict.get("max_map", np.array([]))[iy, ix]) if not np.isnan(map_dict.get("max_map", np.array([]))[iy, ix]) else np.nan,
                    "P25": float(map_dict.get("p25_map", np.array([]))[iy, ix]) if not np.isnan(map_dict.get("p25_map", np.array([]))[iy, ix]) else np.nan,
                    "P75": float(map_dict.get("p75_map", np.array([]))[iy, ix]) if not np.isnan(map_dict.get("p75_map", np.array([]))[iy, ix]) else np.nan,
                    "IQR": float(map_dict.get("iqr_map", np.array([]))[iy, ix]) if not np.isnan(map_dict.get("iqr_map", np.array([]))[iy, ix]) else np.nan,
                    "Count": int(map_dict.get("count_map", np.array([]))[iy, ix]),
                }
            else:
                row = {
                    "X_Value": float(x_val),
                    "Y_Value": float(y_val),
                    "Mean": float(mean_map[iy, ix]) if not np.isnan(mean_map[iy, ix]) else np.nan,
                }
            rows.append(row)
    
    df = pd.DataFrame(rows)
    df.to_csv(output_path, index=False, float_format='%.6f')
    return str(output_path)


def export_map_to_excel(map_dict: Dict[str, Any], output_path: Path, map_name: str = "Map") -> str:
    """
    Export map data to Excel format with multiple sheets:
    - Sheet 1: Mean map (2D table format)
    - Sheet 2: Statistics summary
    - Sheet 3: Detailed per-bin statistics
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to pandas ExcelWriter
        try:
            x_centers = map_dict.get("x_centers", np.array([]))
            y_centers = map_dict.get("y_centers", np.array([]))
            mean_map = map_dict.get("mean_map", np.array([]))
            
            # Sheet 1: Mean map as 2D table
            df_mean = pd.DataFrame(mean_map, index=y_centers, columns=x_centers)
            df_mean.index.name = "Y_Value"
            df_mean.columns.name = "X_Value"
            
            # Sheet 2: Statistics
            stats = map_dict.get("stats", {})
            df_stats = pd.DataFrame([stats])
            
            # Sheet 3: Detailed per-bin
            rows = []
            for iy, y_val in enumerate(y_centers):
                for ix, x_val in enumerate(x_centers):
                    rows.append({
                        "X": float(x_val), "Y": float(y_val),
                        "Mean": float(mean_map[iy, ix]) if not np.isnan(mean_map[iy, ix]) else np.nan,
                        "Median": float(map_dict.get("median_map", np.array([]))[iy, ix]) if not np.isnan(map_dict.get("median_map", np.array([]))[iy, ix]) else np.nan,
                        "StdDev": float(map_dict.get("std_map", np.array([]))[iy, ix]) if not np.isnan(map_dict.get("std_map", np.array([]))[iy, ix]) else np.nan,
                        "Count": int(map_dict.get("count_map", np.array([]))[iy, ix]),
                    })
            df_detail = pd.DataFrame(rows)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_mean.to_excel(writer, sheet_name='Mean Map', index=True)
                df_stats.to_excel(writer, sheet_name='Statistics', index=False)
                df_detail.to_excel(writer, sheet_name='Detailed Data', index=False)
            
            return str(output_path)
        except Exception as e:
            raise MapGeneratorError(f"Excel export failed: {e}")
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Mean Map (2D format)
    ws1 = wb.create_sheet("Mean Map")
    x_centers = map_dict.get("x_centers", np.array([]))
    y_centers = map_dict.get("y_centers", np.array([]))
    mean_map = map_dict.get("mean_map", np.array([]))
    
    # Write headers
    ws1.cell(1, 1, "Y\\X")
    for col_idx, x_val in enumerate(x_centers, start=2):
        ws1.cell(1, col_idx, float(x_val))
    
    # Write data
    for row_idx, y_val in enumerate(y_centers, start=2):
        ws1.cell(row_idx, 1, float(y_val))
        for col_idx, x_val in enumerate(x_centers, start=2):
            val = mean_map[row_idx - 2, col_idx - 2]
            ws1.cell(row_idx, col_idx, float(val) if not np.isnan(val) else "")
    
    # Sheet 2: Statistics
    ws2 = wb.create_sheet("Statistics")
    stats = map_dict.get("stats", {})
    ws2.cell(1, 1, "Metric")
    ws2.cell(1, 2, "Value")
    row = 2
    for key, value in stats.items():
        if isinstance(value, dict):
            ws2.cell(row, 1, key)
            row += 1
            for k, v in value.items():
                ws2.cell(row, 1, f"  {k}")
                ws2.cell(row, 2, v if isinstance(v, (int, float)) else str(v))
                row += 1
        else:
            ws2.cell(row, 1, key)
            ws2.cell(row, 2, value if isinstance(value, (int, float)) else str(value))
            row += 1
    
    wb.save(output_path)
    return str(output_path)


def create_heatmap_trace(map_dict: Dict[str, Any], title: str, z_label: str) -> Optional[go.Figure]:
    # Convert numpy arrays to lists for proper JSON serialization
    x_centers = map_dict.get("x_centers")
    y_centers = map_dict.get("y_centers")
    mean_map = map_dict.get("mean_map")
    
    if x_centers is None or y_centers is None or mean_map is None:
        logger.warning("Missing required data for heatmap: x_centers, y_centers, or mean_map")
        return None
    
    # Convert to lists
    x_list = x_centers.tolist() if isinstance(x_centers, np.ndarray) else list(x_centers)
    y_list = y_centers.tolist() if isinstance(y_centers, np.ndarray) else list(y_centers)
    
    # Convert 2D array to nested list, handling NaN values
    z_list = []
    try:
        if isinstance(mean_map, np.ndarray):
            if mean_map.ndim == 2:
                for row in mean_map:
                    z_row = []
                    for val in row:
                        if np.isnan(val):
                            z_row.append(None)
                        else:
                            try:
                                z_row.append(float(val))
                            except (ValueError, TypeError):
                                z_row.append(None)
                    z_list.append(z_row)
            else:
                # 1D array - shouldn't happen but handle gracefully
                logger.warning(f"mean_map is 1D array, shape: {mean_map.shape}")
                z_list = [[float(val) if not np.isnan(val) else None for val in mean_map.tolist()]]
        else:
            # Not a numpy array - assume list or convert
            z_list = mean_map if isinstance(mean_map, list) else mean_map.tolist()
    except Exception as e:
        logger.error(f"Error converting mean_map to list: {e}", exc_info=True)
        return None
    
    # Validate z_list structure
    if not z_list or len(z_list) == 0:
        logger.warning("z_list is empty - cannot create heatmap")
        return None
    
    if not isinstance(z_list[0], list):
        logger.warning(f"z_list[0] is not a list: {type(z_list[0])} - cannot create 2D heatmap")
        return None
    
    # Validate dimensions
    if len(z_list) != len(y_list):
        logger.warning(f"Z rows ({len(z_list)}) don't match Y length ({len(y_list)})")
    if len(z_list[0]) != len(x_list):
        logger.warning(f"Z cols ({len(z_list[0])}) don't match X length ({len(x_list)})")
    
    # Calculate data bounds for auto-zoom (only where data exists)
    x_data_points = []
    y_data_points = []
    
    for iy, y_val in enumerate(y_list):
        for ix, x_val in enumerate(x_list):
            if iy < len(z_list) and ix < len(z_list[iy]):
                z_val = z_list[iy][ix]
                if z_val is not None and not (isinstance(z_val, float) and np.isnan(z_val)):
                    x_data_points.append(x_val)
                    y_data_points.append(y_val)
    
    # Auto-zoom to data region with 5% padding
    x_range = None
    y_range = None
    if x_data_points and y_data_points:
        x_min, x_max = min(x_data_points), max(x_data_points)
        y_min, y_max = min(y_data_points), max(y_data_points)
        
        # Add 5% padding on each side
        x_padding = (x_max - x_min) * 0.05
        y_padding = (y_max - y_min) * 0.05
        
        x_range = [max(0, x_min - x_padding), x_max + x_padding]
        y_range = [max(0, y_min - y_padding), y_max + y_padding]
        
        logger.debug(f"Auto-zoom: X range {x_range}, Y range {y_range} (from {len(x_data_points)} data points)")
    
    try:
        heatmap = go.Heatmap(
            x=x_list,
            y=y_list,
            z=z_list,
            colorscale="Jet",
            hovertemplate="RPM: %{x:.0f}<br>Torque: %{y:.1f} N·m<br>Value: %{z:.3f}<extra></extra>",
            colorbar=dict(title=z_label),
            hoverongaps=False,
            zsmooth='best',
        )
        fig = go.Figure(data=[heatmap])
        
        # Build layout with auto-zoom
        layout_dict = {
            "title": title,
            "xaxis_title": "RPM",
            "yaxis_title": "Torque (N·m)",
            "template": "plotly_dark",
            "paper_bgcolor": 'black',
            "plot_bgcolor": 'black',
            "font": dict(color='#dce1e6'),
            "height": 500,
            "margin": dict(l=60, r=30, t=50, b=50),
            "autosize": True,
        }
        
        # Apply auto-zoom ranges if calculated
        if x_range and y_range:
            layout_dict["xaxis"] = {"range": x_range}
            layout_dict["yaxis"] = {"range": y_range}
        
        fig.update_layout(**layout_dict)
        logger.debug(f"Created heatmap with dimensions: {len(z_list)}x{len(z_list[0]) if z_list else 0}")
        return fig
    except Exception as e:
        logger.error(f"Failed to create heatmap figure: {e}", exc_info=True)
        return None


def create_validation_plots(map_dict: Dict[str, Any], df_clean: pd.DataFrame, 
                            x_col: str, y_col: str, value_col: str,
                            x_edges: np.ndarray, y_edges: np.ndarray,
                            mean_map: np.ndarray, valid_mask: np.ndarray,
                            x_centers: np.ndarray, y_centers: np.ndarray) -> Dict[str, go.Figure]:
    """
    Create validation plots: scatter (observed vs predicted) and residuals plot.
    MATLAB-level quality analysis.
    """
    validation_plots = {}
    
    try:
        # Get predictions for each data point
        df_clean["x_bin_pred"] = pd.cut(df_clean[x_col], bins=x_edges, labels=False, include_lowest=True)
        df_clean["y_bin_pred"] = pd.cut(df_clean[y_col], bins=y_edges, labels=False, include_lowest=True)
        
        predictions = []
        for _, row in df_clean.iterrows():
            ix = int(row["x_bin_pred"]) if pd.notna(row["x_bin_pred"]) else None
            iy = int(row["y_bin_pred"]) if pd.notna(row["y_bin_pred"]) else None
            if (ix is not None and iy is not None and 
                0 <= ix < len(x_centers) and 0 <= iy < len(y_centers) and 
                valid_mask[iy, ix] and not np.isnan(mean_map[iy, ix])):
                predictions.append(mean_map[iy, ix])
            else:
                predictions.append(np.nan)
        
        df_clean["predicted"] = predictions
        valid_comparison = df_clean[["predicted", value_col]].dropna()
        
        if len(valid_comparison) > 5:
            observed = valid_comparison[value_col].values
            predicted = valid_comparison["predicted"].values
            
            # Scatter plot: Observed vs Predicted
            fig_scatter = go.Figure()
            # Convert numpy arrays to lists for proper JSON serialization
            observed_list = observed.tolist() if isinstance(observed, np.ndarray) else list(observed)
            predicted_list = predicted.tolist() if isinstance(predicted, np.ndarray) else list(predicted)
            
            fig_scatter.add_trace(go.Scatter(
                x=observed_list,
                y=predicted_list,
                mode='markers',
                marker=dict(size=4, opacity=0.6),
                name='Data Points'
            ))
            
            # Add perfect prediction line (y=x)
            min_val = min(np.nanmin(observed), np.nanmin(predicted))
            max_val = max(np.nanmax(observed), np.nanmax(predicted))
            fig_scatter.add_trace(go.Scatter(
                x=[min_val, max_val],
                y=[min_val, max_val],
                mode='lines',
                line=dict(color='red', dash='dash', width=2),
                name='Perfect Prediction'
            ))
            
            fig_scatter.update_layout(
                title="Observed vs Predicted",
                xaxis_title="Observed Value",
                yaxis_title="Predicted Value",
                template="plotly_white",
                width=500,
                height=500
            )
            validation_plots["scatter_observed_vs_predicted"] = fig_scatter
            
            # Residuals plot
            residuals = observed - predicted
            fig_residuals = go.Figure()
            # Convert numpy arrays to lists for proper JSON serialization
            predicted_list = predicted.tolist() if isinstance(predicted, np.ndarray) else list(predicted)
            residuals_list = residuals.tolist() if isinstance(residuals, np.ndarray) else list(residuals)
            
            fig_residuals.add_trace(go.Scatter(
                x=predicted_list,
                y=residuals_list,
                mode='markers',
                marker=dict(size=4, opacity=0.6),
                name='Residuals'
            ))
            
            # Add zero line
            fig_residuals.add_trace(go.Scatter(
                x=[min(predicted), max(predicted)],
                y=[0, 0],
                mode='lines',
                line=dict(color='red', dash='dash', width=2),
                name='Zero Residual'
            ))
            
            fig_residuals.update_layout(
                title="Residuals Plot",
                xaxis_title="Predicted Value",
                yaxis_title="Residual (Observed - Predicted)",
                template="plotly_white",
                width=500,
                height=500
            )
            validation_plots["residuals_plot"] = fig_residuals
            
            # Residuals histogram
            fig_hist = go.Figure()
            # Convert numpy array to list for proper JSON serialization
            residuals_list = residuals.tolist() if isinstance(residuals, np.ndarray) else list(residuals)
            fig_hist.add_trace(go.Histogram(
                x=residuals_list,
                nbinsx=30,
                name='Residuals Distribution'
            ))
            fig_hist.update_layout(
                title="Residuals Distribution",
                xaxis_title="Residual Value",
                yaxis_title="Frequency",
                template="plotly_white",
                width=500,
                height=400
            )
            validation_plots["residuals_histogram"] = fig_hist
            
    except Exception as e:
        logger.warning("Validation plots generation failed: %s", e)
    
    return validation_plots


def create_surface_trace(map_dict: Dict[str, Any], title: str, z_label: str) -> Optional[go.Figure]:
    surface = map_dict.get("surface_data")
    if not surface:
        logger.warning("No surface_data in map_dict")
        return None
    
    z_data = surface.get("z")
    if z_data is None:
        logger.warning("surface_data.z is None")
        return None
    
    # Check if z_data is empty or all NaN
    if isinstance(z_data, np.ndarray):
        if z_data.size == 0 or np.all(np.isnan(z_data)):
            logger.warning("Surface z_data is empty or all NaN")
            # Still create the plot but it might not display - let Plotly handle it
            pass
    
    # Convert numpy arrays to lists for proper JSON serialization
    x_data = surface.get("x")
    y_data = surface.get("y")
    
    if x_data is None or y_data is None:
        logger.warning("Missing x or y data in surface_data")
        return None
    
    # Convert to lists (handling 2D arrays)
    if isinstance(x_data, np.ndarray):
        if x_data.ndim == 2:
            x_list = [[float(x) if not np.isnan(x) else None for x in row] for row in x_data]
        else:
            x_list = [float(x) if not np.isnan(x) else None for x in x_data.tolist()]
    else:
        x_list = [float(x) if not (isinstance(x, float) and np.isnan(x)) else None for x in list(x_data)]
    
    if isinstance(y_data, np.ndarray):
        if y_data.ndim == 2:
            y_list = [[float(y) if not np.isnan(y) else None for y in row] for row in y_data]
        else:
            y_list = [float(y) if not np.isnan(y) else None for y in y_data.tolist()]
    else:
        y_list = [float(y) if not (isinstance(y, float) and np.isnan(y)) else None for y in list(y_data)]
    
    # Convert z (2D array) to nested list, handling NaN values properly
    z_list = []
    try:
        if isinstance(z_data, np.ndarray):
            if z_data.ndim == 2:
                for row in z_data:
                    z_row = []
                    for val in row:
                        if np.isnan(val):
                            z_row.append(None)
                        else:
                            try:
                                z_row.append(float(val))
                            except (ValueError, TypeError):
                                z_row.append(None)
                    z_list.append(z_row)
            else:
                # 1D array - convert to 2D or handle appropriately
                z_list = [[float(val) if not np.isnan(val) else None for val in z_data.tolist()]]
        else:
            # Not a numpy array - assume it's already a list or iterable
            z_list = z_data if isinstance(z_data, list) else list(z_data)
    except Exception as e:
        logger.error(f"Error converting z_data to list: {e}")
        return None
    
    # Validate dimensions match
    if len(x_list) > 0 and len(y_list) > 0 and len(z_list) > 0:
        if isinstance(x_list[0], list) and isinstance(y_list[0], list) and isinstance(z_list[0], list):
            x_rows, x_cols = len(x_list), len(x_list[0])
            y_rows, y_cols = len(y_list), len(y_list[0])
            z_rows, z_cols = len(z_list), len(z_list[0])
            
            if x_rows != y_rows or x_cols != y_cols:
                logger.warning(f"X and Y dimensions don't match: X({x_rows}x{x_cols}) vs Y({y_rows}x{y_cols})")
            if z_rows != x_rows or z_cols != x_cols:
                logger.warning(f"Z dimensions don't match X/Y: Z({z_rows}x{z_cols}) vs X/Y({x_rows}x{x_cols})")
                # Try to fix dimension mismatch by reshaping if possible
                if z_rows * z_cols == x_rows * x_cols:
                    logger.info("Attempting to reshape z_data to match x/y dimensions")
                    try:
                        z_list = [z_list[i * z_cols:(i + 1) * z_cols] for i in range(x_rows)]
                    except Exception:
                        pass
    
    # Validate that we have valid data
    if not z_list or len(z_list) == 0:
        logger.warning("z_list is empty - cannot create surface plot")
        return None
    
    # Check if all data is None/NaN
    all_nan = True
    for row in z_list:
        if isinstance(row, list):
            if any(v is not None for v in row):
                all_nan = False
                break
        elif row is not None:
            all_nan = False
            break
    
    if all_nan:
        logger.warning("All z_data values are NaN - surface plot may not display")
    
    try:
        # Determine axis labels from title or use defaults
        x_label = "RPM" if "rpm" in title.lower() or "engine" in title.lower() else "X"
        y_label = "Torque (N·m)" if "torque" in title.lower() or "engine" in title.lower() else "Y"
        
        # Calculate data bounds for auto-zoom (3D)
        x_values = []
        y_values = []
        z_values = []
        
        # Extract valid data points
        if isinstance(x_list[0], list) and isinstance(y_list[0], list) and isinstance(z_list[0], list):
            # 2D meshgrid format
            for i, z_row in enumerate(z_list):
                if i < len(x_list) and i < len(y_list):
                    for j, z_val in enumerate(z_row):
                        if j < len(x_list[i]) and j < len(y_list[i]):
                            if z_val is not None and not (isinstance(z_val, float) and np.isnan(z_val)):
                                x_values.append(x_list[i][j])
                                y_values.append(y_list[i][j])
                                z_values.append(z_val)
        else:
            # 1D format - try to extract
            for i, z_row in enumerate(z_list):
                if isinstance(z_row, list):
                    for j, z_val in enumerate(z_row):
                        if z_val is not None and not (isinstance(z_val, float) and np.isnan(z_val)):
                            if i < len(x_list) and j < len(x_list[i] if isinstance(x_list[0], list) else x_list):
                                x_val = x_list[i][j] if isinstance(x_list[0], list) else x_list[j]
                                y_val = y_list[i][j] if isinstance(y_list[0], list) else y_list[i]
                                x_values.append(x_val)
                                y_values.append(y_val)
                                z_values.append(z_val)
        
        # Calculate auto-zoom ranges with 5% padding
        scene_ranges = {}
        if x_values and y_values and z_values:
            x_min, x_max = min(x_values), max(x_values)
            y_min, y_max = min(y_values), max(y_values)
            z_min, z_max = min(z_values), max(z_values)
            
            x_padding = (x_max - x_min) * 0.05 if x_max > x_min else 0
            y_padding = (y_max - y_min) * 0.05 if y_max > y_min else 0
            z_padding = (z_max - z_min) * 0.05 if z_max > z_min else 0
            
            scene_ranges = {
                "xaxis": {"range": [max(0, x_min - x_padding), x_max + x_padding]},
                "yaxis": {"range": [max(0, y_min - y_padding), y_max + y_padding]},
                "zaxis": {"range": [z_min - z_padding, z_max + z_padding]},
            }
            logger.debug(f"Auto-zoom 3D: X={scene_ranges['xaxis']['range']}, Y={scene_ranges['yaxis']['range']}, Z={scene_ranges['zaxis']['range']}")
        
        fig = go.Figure(
            data=[
                go.Surface(
                    x=x_list,
                    y=y_list,
                    z=z_list,
                    colorscale="Jet",
                    colorbar=dict(title=z_label),
                    hovertemplate=f"{x_label}: %{{x:.0f}}<br>{y_label}: %{{y:.1f}}<br>{z_label}: %{{z:.3f}}<extra></extra>",
                    showscale=True,
                    connectgaps=False,  # Don't connect gaps for better visualization
                )
            ]
        )
        
        # Build scene dict with auto-zoom
        scene_dict = {
            "xaxis_title": x_label,
            "yaxis_title": y_label, 
            "zaxis_title": z_label,
            "camera": dict(eye=dict(x=1.5, y=1.5, z=1.2))
        }
        
        # Apply auto-zoom ranges if calculated
        if scene_ranges:
            scene_dict["xaxis"] = scene_ranges["xaxis"]
            scene_dict["yaxis"] = scene_ranges["yaxis"]
            scene_dict["zaxis"] = scene_ranges["zaxis"]
        
        fig.update_layout(
            title=title,
            scene=scene_dict,
            template="plotly_dark",
            height=600,
            margin=dict(l=0, r=0, t=30, b=0),
        )
        logger.debug(f"Created surface plot with {len(z_list)} rows, {len(z_list[0]) if z_list else 0} cols")
        return fig
    except Exception as e:
        logger.error(f"Failed to create surface plot: {e}", exc_info=True)
        return None


def read_mdf_to_dataframe(path: Union[str, Path], channels: Optional[List[str]] = None) -> Tuple[pd.DataFrame, Dict[str, List[str]]]:
    if MDF is None:
        raise MapGeneratorError("asammdf not available to read MDF/MF4 files")
    p = Path(path)
    logger.info("Reading MDF file %s (asammdf v%s)", p.name, ASAMMDF_VERSION)
    ambiguous: Dict[str, List[str]] = {}
    try:
        mdf = MDF(str(p))
    except Exception as exc:  # pragma: no cover - file specific issues
        raise MapGeneratorError(f"Failed to open MDF file {path}: {exc}")

    # Use advanced signal mapping to find actual signals first
    if channels is None and find_signal_by_role:
        # Find signals by role using advanced mapping
        available_channels = list(mdf.channels_db.keys())
        signal_roles = list(REQUIRED_SIGNALS.keys())
        
        # Map roles to actual channel names
        role_to_channel = {}
        for role in signal_roles:
            # Try to find signal by role first
            found = find_signal_by_role(mdf, role)
            if not found:
                # Map internal role names to signal_mapping role names
                role_mapping = {
                    "rpm": "rpm",
                    "torque": "torque",
                    "lambda_raw": "lambda",
                    "intake_air_temp_c": "intake_air_temp",
                    "coolant_temp": "coolant_temp",
                    "exhaust_temp": "exhaust_temp",
                    "map_sensor": "map_sensor",
                    "fuel_vol_consumption": "fuel_rate",
                    "air_mass_flow": "air_mass_flow",
                    "oil_temp": "oil_temp",
                    "batt_voltage": "battery_voltage",
                }
                mapped_role = role_mapping.get(role, role)
                found = find_signal_by_role(mdf, mapped_role)
            
            if found:
                role_to_channel[role] = found
        
        # Use found channels plus any additional requested
        requested = list(role_to_channel.values()) if role_to_channel else []
        if not requested:
            # Fallback: use first few candidates from each role (not all 622!)
            requested = []
            for role in signal_roles:
                candidates = REQUIRED_SIGNALS.get(role, [])[:5]  # Only try first 5 per role
                requested.extend(candidates)
    else:
        requested = channels or list({alias for aliases in REQUIRED_SIGNALS.values() for alias in aliases})
    
    data_cols: Dict[str, pd.Series] = {}

    for channel in requested:
        try:
            series = mdf.get(channel)
        except MdfException as exc:
            msg = str(exc)
            if "Multiple occurrences" in msg and hasattr(mdf, "whereis"):
                occurrences = mdf.whereis(channel)
                if occurrences:
                    ambiguous[channel] = [f"group={grp}, index={idx}" for grp, idx in occurrences]
                    grp, idx = occurrences[0]
                    series = mdf.get(group=int(grp), index=int(idx))
                else:
                    logger.warning("Channel '%s' not found", channel)
                    continue
            else:
                logger.warning("Failed to read channel '%s': %s", channel, exc)
                continue
        except Exception as exc:
            logger.warning("Error extracting channel '%s': %s", channel, exc)
            continue

        samples = pd.Series(series.samples, name=channel)
        if samples.dropna().empty:
            logger.debug("Channel '%s' empty after dropna", channel)
            continue
        data_cols[channel] = samples.reset_index(drop=True)

    if not data_cols:
        raise MapGeneratorError(f"No requested channels found in {path}.")

    df = pd.concat(data_cols, axis=1)
    df.ffill(inplace=True)
    df.bfill(inplace=True)

    if df.dropna(how="all").empty:
        raise MapGeneratorError(f"All data from {path} became NaN after filling")

    return df, ambiguous


def read_csv_to_dataframe(path: Path) -> pd.DataFrame:
    encodings = ["utf-8", "latin-1", "iso-8859-1", "windows-1252"]
    last_exc: Optional[Exception] = None
    for enc in encodings:
        try:
            # Read first few rows to detect structure (without header first)
            sample_no_header = pd.read_csv(path, encoding=enc, nrows=4, header=None, low_memory=False)
            
            # Detect pattern: headers in row 0, descriptions in row 1, units in row 2, data from row 3
            skip_rows = 0
            if len(sample_no_header) >= 3:
                first_col = sample_no_header.iloc[:, 0]
                
                # Pattern: Row 0=headers, Row 1=descriptions (text), Row 2=units (short text), Row 3=data (numeric)
                row1_first = str(first_col.iloc[1] if len(first_col) > 1 else '').strip()
                row2_first = str(first_col.iloc[2] if len(first_col) > 2 else '').strip()
                row3_first = str(first_col.iloc[3] if len(first_col) > 3 else '').strip()
                
                # Try to convert row 3 to numeric to see if it's data
                is_row3_numeric = False
                try:
                    float(str(row3_first).replace(',', '').strip())
                    is_row3_numeric = True
                except (ValueError, TypeError):
                    pass
                
                # If row 3 is numeric and rows 1-2 are text (likely descriptions/units), skip 2 rows
                if (is_row3_numeric and 
                    len(row1_first) > 3 and  # Row 1 is descriptive text (not just a number)
                    len(row2_first) < 50):   # Row 2 is likely units (short)
                    skip_rows = 2  # Skip description and units rows
                # Pattern: Row 0=headers, Row 1=descriptions, Row 2=data
                elif (is_row3_numeric and len(row1_first) > 3):
                    # Check if row 2 could be units (very short)
                    if len(row2_first) < 20:
                        skip_rows = 2
                    else:
                        skip_rows = 1
            
            # Read full file - header=0 means use first row as headers
            # If skip_rows > 0, skip those rows after header
            if skip_rows > 0:
                skip_list = list(range(1, skip_rows + 1))
                df = pd.read_csv(path, encoding=enc, header=0, skiprows=skip_list, low_memory=False)
            else:
                df = pd.read_csv(path, encoding=enc, header=0, low_memory=False)
            
            # Clean up: try to convert columns to numeric where possible
            for col in df.columns:
                try:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                except Exception:
                    pass
            
            return df
        except (UnicodeDecodeError, UnicodeError) as exc:
            last_exc = exc
            continue
        except Exception as exc:
            # If skiprows detection fails, try without skipping
            try:
                return pd.read_csv(path, encoding=enc, header=0, low_memory=False)
            except Exception:
                last_exc = exc
                continue
    raise last_exc or IOError(f"Unable to read CSV file {path}")


def parse_bins(spec: Optional[Union[str, List[float], np.ndarray]]) -> Optional[np.ndarray]:
    if spec is None:
        return None
    if isinstance(spec, (list, np.ndarray)):
        arr = np.asarray(spec, dtype=float)
        # Ensure sorted and unique
        arr = np.unique(np.sort(arr))
        if len(arr) < 2:
            return None
        return arr
    if isinstance(spec, str):
        spec = spec.strip()
        if ":" in spec:
            parts = [float(x) for x in spec.split(":" )]
            if len(parts) == 3:
                start, stop, step = parts
                if step <= 0:
                    return None
                # Ensure start < stop
                if start > stop:
                    start, stop = stop, start
                arr = np.arange(start, stop + step * 0.9999, step)
                return np.unique(np.sort(arr))
        values = [float(x.strip()) for x in spec.split(",") if x.strip()]
        if len(values) < 2:
            return None
        arr = np.asarray(values, dtype=float)
        return np.unique(np.sort(arr))
    try:
        arr = np.asarray(spec, dtype=float)
        arr = np.unique(np.sort(arr))
        if len(arr) < 2:
            return None
        return arr
    except Exception:  # pragma: no cover - malformed spec
        return None


def summarize_map(map_dict: Dict[str, Any]) -> Dict[str, Any]:
    """
    Enhanced summary with MATLAB-level statistics.
    """
    stats = map_dict.get("stats", {})
    meta = map_dict.get("meta", {})
    quality = stats.get("quality_metrics", {})
    
    return {
        "cells_total": int(meta.get("total_cells", 0)),
        "cells_filled": int(meta.get("cells_with_data", 0)),
        "coverage_pct": float(stats.get("coverage_ratio", 0) * 100.0),
        "mean": stats.get("global_mean"),
        "median": stats.get("global_median"),
        "min": stats.get("global_min"),
        "max": stats.get("global_max"),
        "std": stats.get("global_std"),
        "p25": stats.get("global_p25"),
        "p75": stats.get("global_p75"),
        "quality_r_squared": quality.get("r_squared"),
        "quality_rmse": quality.get("rmse"),
        "quality_mae": quality.get("mae"),
        "quality_mape": quality.get("mape"),
        "validation_samples": quality.get("n_samples_validated"),
    }


def generate_maps_from_df(
    df: pd.DataFrame,
    rpm_bins: Optional[np.ndarray] = None,
    tq_bins: Optional[np.ndarray] = None,
    min_samples: int = MIN_SAMPLES_PER_BIN,
    output_format: str = "plotly_json",
    interp_method: str = "linear",
    smoothing: float = 0.5,
    contour_levels: int = 10,
    enable_contours: bool = True,
    enable_surface: bool = True,
    overrides: Optional[Dict[str, str]] = None,
    preset_maps: Optional[List[str]] = None,
    filter_steady_state: bool = False,
    filter_outliers: bool = False,
) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, str]], Dict[str, Any], List[Dict[str, Any]], Dict[str, List[str]]]:
    df, mapping, report_rows, missing_signals = derive_signals(df, overrides)
    quality_report = validate_data_quality(df, ["rpm", "torque"])

    if any(sig not in df.columns for sig in CRITICAL_SIGNALS):
        missing = [sig for sig in CRITICAL_SIGNALS if sig not in df.columns]
        raise MapGeneratorError(f"Missing required signals for map generation: {missing}")

    if missing_signals:
        logger.info("Proceeding with optional missing signals: %s", list(missing_signals.keys()))

    # Fix for numpy array boolean check: use None check instead of truthiness
    if rpm_bins is None:
        rpm_bins = DEFAULT_RPM_BINS
    if tq_bins is None:
        tq_bins = DEFAULT_TQ_BINS

    map_configs = [
        ("engine_bsfc", "bsfc_gpkwh", "Engine BSFC Map", "g/kWh"),
        ("exhaust_temperature", "exhaust_temp", "Exhaust Temperature", "°C"),
        ("air_fuel_ratio", "lambda_raw", "Air-Fuel Ratio (λ)", "λ"),
        ("motor_efficiency", "efficiency", "Motor Efficiency", "%"),
        ("volumetric_efficiency", "volumetric_efficiency", "Volumetric Efficiency", "%"),
        ("thermal_efficiency", "thermal_efficiency", "Thermal Efficiency", "%"),
        ("bmep", "bmep_kpa", "BMEP (Brake Mean Effective Pressure)", "kPa"),
        ("mean_piston_speed", "mean_piston_speed_ms", "Mean Piston Speed", "m/s"),
    ]

    if preset_maps:
        map_configs = [cfg for cfg in map_configs if cfg[0] in preset_maps]

    summaries: List[Dict[str, Any]] = []
    plots: Dict[str, Dict[str, str]] = {}
    
    for name, column, title, label in map_configs:
        if column not in df.columns or not df[column].notna().any():
            continue
        try:
            map_data = create_calibration_map(
                df,
                "rpm",
                "torque",
                column,
                rpm_bins,
                tq_bins,
                min_samples,
                interp_method,
                (100, 100),
                smoothing,
                contour_levels,
                enable_contours,
                enable_surface,
                filter_steady_state,
                filter_outliers,
            )
            summaries.append({"map": name, **summarize_map(map_data)})
            if output_format == "plotly_json":
                # Create heatmap plot
                try:
                    heatmap_fig = create_heatmap_trace(map_data, title, label)
                    if heatmap_fig is not None:
                        heatmap_json = heatmap_fig.to_json()
                        # Use consistent structure with other plots
                        plots[f"{name}_heatmap"] = {
                            "type": "plotly",
                            "plotly_json": heatmap_json
                        }
                        logger.info(f"✅ Created heatmap plot for {name} - {len(heatmap_json)} chars")
                    else:
                        logger.warning(f"⚠️  create_heatmap_trace returned None for {name}")
                except Exception as e:
                    logger.error(f"❌ Failed to create heatmap plot for {name}: {e}", exc_info=True)
                
                # Create surface plot
                try:
                    surface_fig = create_surface_trace(map_data, f"{title} (Surface)", label)
                    if surface_fig is not None:
                        surface_json = surface_fig.to_json()
                        # Use consistent structure with other plots
                        plots[f"{name}_surface"] = {
                            "type": "plotly",
                            "plotly_json": surface_json
                        }
                        logger.info(f"✅ Created surface plot for {name} - {len(surface_json)} chars")
                    else:
                        logger.warning(f"⚠️  create_surface_trace returned None for {name}")
                        # Debug why surface_fig is None
                        if map_data.get("surface_data"):
                            surface_data = map_data.get("surface_data")
                            logger.warning(f"  - surface_data exists: {surface_data is not None}")
                            if surface_data:
                                logger.warning(f"  - surface_data has z: {surface_data.get('z') is not None}")
                                z_data = surface_data.get('z')
                                if z_data is not None and isinstance(z_data, np.ndarray):
                                    logger.warning(f"  - z_data shape: {z_data.shape}, valid points: {np.sum(~np.isnan(z_data))}")
                        else:
                            logger.warning(f"  - No surface_data in map_data for {name}")
                except Exception as e:
                    logger.error(f"❌ Failed to create surface plot for {name}: {e}", exc_info=True)
                
                # Add validation plots if available (MATLAB-level quality analysis)
                validation_plots = map_data.get("validation_plots", {})
                for plot_name, plot_data in validation_plots.items():
                    plots[f"{name}_{plot_name}"] = plot_data
        except Exception as exc:
            logger.warning("Failed to generate map '%s': %s", name, exc)

    return summaries, plots, quality_report, report_rows, missing_signals


def compute_map(
    files: List[Path],
    rpm_bins: Optional[np.ndarray] = None,
    tq_bins: Optional[np.ndarray] = None,
    min_samples_per_bin: int = MIN_SAMPLES_PER_BIN,
    output_format: str = "plotly_json",
    interp_method: str = "linear",
    smoothing: float = 0.5,
    contour_levels: int = 10,
    enable_contours: bool = True,
    enable_surface: bool = True,
    overrides: Optional[Dict[str, str]] = None,
               map_type: Optional[str] = None,
    outdir: Optional[Path] = None,
    preset: Optional[str] = None,
    progress_callback: Optional[callable] = None
) -> Dict[str, Any]:
    def update_progress(stage, progress, message):
        if progress_callback:
            progress_callback(stage, progress, message)

    start_time = time.time()
    preset_maps = None
    filter_steady_state = False
    filter_outliers = False

    if preset:
        preset_cfg = PRESET_TEMPLATES.get(preset)
        if preset_cfg:
            logger.info("Applying preset '%s': %s", preset, preset_cfg.get("label"))
            if rpm_bins is None and preset_cfg.get("x_bins"):
                rpm_bins = parse_bins(preset_cfg["x_bins"])
            if tq_bins is None and preset_cfg.get("y_bins"):
                tq_bins = parse_bins(preset_cfg["y_bins"])
            if min_samples_per_bin == MIN_SAMPLES_PER_BIN and preset_cfg.get("min_samples_per_bin"):
                min_samples_per_bin = int(preset_cfg["min_samples_per_bin"])
            if interp_method == "linear" and preset_cfg.get("interp_method"):
                interp_method = preset_cfg["interp_method"]
            if smoothing == 0.5 and preset_cfg.get("smoothing") is not None:
                smoothing = float(preset_cfg["smoothing"])
            preset_maps = preset_cfg.get("maps")
            # Apply advanced filtering from preset
            filter_steady_state = preset_cfg.get("filter_steady_state", False)
            filter_outliers = preset_cfg.get("filter_outliers", False)
        else:
            logger.warning("Preset '%s' not found", preset)

    update_progress("file_loading", 0, "Starting file processing...")
    aggregated_frames: List[pd.DataFrame] = []
    total_rows = 0
    problems: List[Dict[str, Any]] = []
    loaded_files: List[str] = []
    ambiguous_by_file: Dict[str, Dict[str, List[str]]] = {}

    total_files = len(files)
    for file_idx, file_path in enumerate(files):
        path = Path(file_path)
        if not path.exists():
            problems.append({"file": str(path), "error": "not_found"})
            continue
        
        # Update progress during file loading
        file_progress = int((file_idx / total_files) * 80) + 10  # 10-90%
        update_progress("file_loading", file_progress, f"Loading file {file_idx + 1}/{total_files}: {path.name}...")
            
        try:
            file_size = path.stat().st_size / (1024 * 1024)
            suffix = path.suffix.lower()
            df: Optional[pd.DataFrame] = None
            ambiguous: Dict[str, List[str]] = {}

            if suffix in {".mf4", ".mf3", ".mdf"}:
                if MDF is None:
                    problems.append({"file": str(path), "error": "asammdf_missing"})
                    continue
                if file_size > MAX_FILE_SIZE_MB:
                    logger.info("Large MDF detected (%.2f MB); chunked reading", file_size)
                    with MDF(str(path), memory="memmap") as mdf_file:  # type: ignore[attr-defined]
                        mdf_file.configure(raise_on_multiple_occurrences=False)
                        filtered = mdf_file.filter(list(REQUIRED_SIGNALS.keys()))
                        iterator = filtered.iter_to_dataframe(chunk_size=CHUNK_SIZE, use_interpolation=True)
                        chunks = [chunk for chunk in iterator if chunk is not None and not chunk.empty]
                    if chunks:
                        df = pd.concat(chunks, ignore_index=True)
                else:
                    df, ambiguous = read_mdf_to_dataframe(path)
            elif suffix == ".csv":
                df = read_csv_to_dataframe(path)
            elif suffix in {".xlsx", ".xls"}:
                try:
                    if suffix == ".xlsx":
                        df = pd.read_excel(path, engine='openpyxl')
                    else:
                        # .xls files - try openpyxl first, then default
                        try:
                            df = pd.read_excel(path, engine='openpyxl')
                        except Exception:
                            df = pd.read_excel(path)
                    # Clean up: try to convert columns to numeric where possible
                    for col in df.columns:
                        try:
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                        except Exception:
                            pass
                except Exception as exc:
                    problems.append({"file": str(path), "error": f"excel_read_error: {exc}"})
                    continue
            else:
                problems.append({"file": str(path), "error": "unsupported_file_type"})
                continue

            if df is None or df.empty:
                problems.append({"file": str(path), "error": "no_rows_read"})
                continue

            if ambiguous:
                ambiguous_by_file[path.name] = ambiguous

            loaded_files.append(str(path))
            total_rows += len(df)
            aggregated_frames.append(df)
        except Exception as exc:  # pragma: no cover - file-specific errors
            problems.append({"file": str(path), "error": f"{type(exc).__name__}: {exc}"})
            logger.error("Error processing file %s: %s", path, exc)
            continue

    update_progress("file_loading", 100, f"Loaded {len(aggregated_frames)} files.")

    if not aggregated_frames:
        return {
            "tables": {"Map Summary": [], "Signal Mapping": []},
            "plots": {},
            "meta": {
                "ok": False,
                "error": "no_input_data",
                "problems": problems,
                "files": loaded_files,
            },
        }

    update_progress("signal_mapping", 10, "Merging dataframes...")
    df_all = pd.concat(aggregated_frames, ignore_index=True, sort=False)
    logger.info("Merged %d files into %d samples", len(aggregated_frames), len(df_all))
    update_progress("signal_mapping", 50, "Deriving signals...")

    try:
        summary_rows, plots, quality_report, mapping_report, missing_signals = generate_maps_from_df(
            df_all,
            rpm_bins,
            tq_bins,
            min_samples_per_bin,
            output_format,
            interp_method,
            smoothing,
            contour_levels,
            enable_contours,
            enable_surface,
            overrides,
            preset_maps,
            filter_steady_state,
            filter_outliers,
        )
        update_progress("signal_mapping", 70, "Signal mapping in progress...")
        update_progress("data_validation", 50, "Validating data quality...")
        update_progress("physics_calculations", 30, "Calculating derived signals...")
        update_progress("map_generation", 20, "Generating maps...")
        update_progress("signal_mapping", 100, "Signal mapping complete.")
        update_progress("data_validation", 100, "Validation complete.")
        update_progress("physics_calculations", 100, "Physics calculations complete.")
        update_progress("map_generation", 100, "Map generation complete.")
    except Exception as exc:
        logger.error("Map generation failed: %s", traceback.format_exc())
        update_progress("map_generation", 0, f"Error: {str(exc)[:100]}")
        return {
            "tables": {"Map Summary": [], "Signal Mapping": []},
            "plots": {},
            "meta": {"ok": False, "error": "map_generation_failed", "detail": str(exc)},
        }

    processing_time = round(time.time() - start_time, 2)
    update_progress("visualization", 100, "Visualization complete.")
    meta = {
        "ok": True,
        "files": loaded_files,
        "rows": int(total_rows),
        "problems": problems,
        "timestamp": datetime.now().isoformat(),
        "processing_time_sec": processing_time,
        "data_quality": quality_report,
        "signal_mapping": mapping_report,
        "ambiguous_channels": ambiguous_by_file,
        "missing_signals": missing_signals,
        "settings": {
            "preset": preset,
            "min_samples_per_bin": int(min_samples_per_bin),
            "interp_method": interp_method,
            "smoothing": float(smoothing),
            "contour_levels": contour_levels,
            "overrides": overrides or {},
        },
    }

    # Sample data payload for UI previews
    numeric_df = df_all.select_dtypes(include=[np.number])
    if numeric_df.empty:
        numeric_df = df_all
    max_rows = 3000
    if len(numeric_df) > max_rows:
        indices = np.linspace(0, len(numeric_df) - 1, max_rows, dtype=int)
        export_df = numeric_df.iloc[indices].reset_index(drop=True)
    else:
        export_df = numeric_df.reset_index(drop=True)

    samples: Dict[str, Any] = {
        "rows_exported": int(len(export_df)),
        "rows_total": int(len(numeric_df)),
        "columns": list(export_df.columns),
        "samples": {col: [None if pd.isna(val) else (float(val) if isinstance(val, (np.integer, np.floating)) else val) for val in export_df[col].tolist()] for col in export_df.columns},
        "stats": {
            col: {
                "min": float(export_df[col].min()) if export_df[col].notna().any() else None,
                "max": float(export_df[col].max()) if export_df[col].notna().any() else None,
                "median": float(export_df[col].median()) if export_df[col].notna().any() else None,
                "count": int(export_df[col].notna().sum()),
            }
            for col in export_df.columns
        },
    }

    result = {
        "tables": {"Map Summary": summary_rows, "Signal Mapping": mapping_report},
        "plots": plots,
        "meta": meta,
        "samples": samples,
    }

    if outdir is not None:
        outdir.mkdir(parents=True, exist_ok=True)
        output_path = outdir / "map_output.json"
        try:
            # Use make_json_serializable to handle NaN values properly
            try:
                from app import make_json_serializable
                serializable_result = make_json_serializable(result)
                with open(output_path, "w", encoding="utf-8") as handle:
                    json.dump(serializable_result, handle, indent=2, ensure_ascii=False)
            except ImportError:
                # Fallback: use json.dump with default handler (may still have NaN issues)
                with open(output_path, "w", encoding="utf-8") as handle:
                    json.dump(result, handle, indent=2, ensure_ascii=False, default=str)
            logger.info("Wrote map_output.json to %s", output_path)
        except Exception as e:
            logger.warning("Failed to write map_output.json: %s", e)

    return result


def compute_map_plotly(files: List[str], **kwargs: Any) -> Dict[str, Any]:
    paths = [Path(f) for f in files]
    min_samples = kwargs.get("min_samples_per_bin", kwargs.get("min_samples", MIN_SAMPLES_PER_BIN))
    return compute_map(
        paths, 
        rpm_bins=kwargs.get("rpm_bins"),
        tq_bins=kwargs.get("tq_bins"),
        min_samples_per_bin=int(min_samples),
        output_format=kwargs.get("output_format", "plotly_json"),
        interp_method=kwargs.get("interp_method", "linear"),
        smoothing=float(kwargs.get("smoothing", 0.5)),
        contour_levels=int(kwargs.get("contour_levels", 10)),
        enable_contours=bool(kwargs.get("enable_contours", True)),
        enable_surface=bool(kwargs.get("enable_surface", True)),
        overrides=kwargs.get("overrides"),
        outdir=kwargs.get("outdir"),
        preset=kwargs.get("preset"),
        progress_callback=kwargs.get("progress_callback"),
    )


# CLI -----------------------------------------------------------------------
def cli_main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Empirical map generator")
    parser.add_argument("--files", nargs="+", required=True, help="MDF/MF4/CSV files")
    parser.add_argument("--rpm-bins", type=str, help="RPM bins like '0:8000:250'")
    parser.add_argument("--tq-bins", type=str, help="Torque bins like '-200:2000:10'")
    parser.add_argument("--min-samples", type=int, default=MIN_SAMPLES_PER_BIN)
    parser.add_argument("--smoothing", type=float, default=0.5)
    parser.add_argument("--outdir", type=str, help="Output directory for JSON results")
    parser.add_argument("--preset", type=str, help="Preset key (e.g. ci_engine_default)")
    args = parser.parse_args(argv)

    rpm_bins = parse_bins(args.rpm_bins)
    tq_bins = parse_bins(args.tq_bins)
    outdir = Path(args.outdir) if args.outdir else Path.cwd() / "maps_outputs" / f"run_{uuid.uuid4().hex[:8]}"

    try:
        result = compute_map_plotly(
            args.files,
            rpm_bins=rpm_bins,
            tq_bins=tq_bins,
            min_samples_per_bin=args.min_samples,
            smoothing=args.smoothing,
            outdir=outdir,
            preset=args.preset,
        )
    except Exception as exc:
        logger.error("Failed to generate maps: %s", exc)
        print(json.dumps({"error": str(exc)}, indent=2))
        return 1
        
    print(json.dumps(result.get("meta", {}), indent=2))
    print(f"Map generation complete. Output written to {outdir / 'map_output.json'}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(cli_main()) 